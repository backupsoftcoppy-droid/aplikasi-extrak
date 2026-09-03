import io
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pdfplumber
import streamlit as st

# ==========================================
# METODE OTENTIKASI STREAMLIT
# ==========================================
def check_password():
    def password_entered():
        if st.session_state["password"] == st.secrets.get("APP_PASSWORD", "deon22"):
            st.session_state["password_correct"] = True
            del st.session_state["password"]
        else:
            st.session_state["password_correct"] = False

    if st.session_state.get("password_correct", False):
        return True

    st.set_page_config(page_title="Login - SPX Processor", page_icon="🔒", layout="centered")
    st.title("🔒 Akses Terbatas - SPX Internal")
    st.write("Silakan masukkan password tim internal untuk melanjutkan.")

    st.text_input("Masukkan Password", type="password", on_change=password_entered, key="password")
    
    if "password_correct" in st.session_state and not st.session_state["password_correct"]:
        st.error("❌ Password salah. Silakan coba lagi.")
        
    return False


# ==========================================
# FUNGSIONALITAS: PROSES PDF LAPORAN SCAN
# ==========================================
def process_pdf_scan(pdf_bytes):
    trips_summary = []
    to_details = []

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        full_text = "\n".join([page.extract_text() or '' for page in pdf.pages])

    docs = full_text.split("Surat Jalan Line Haul")

    for doc in docs[1:]:
        def get_val(pattern, text):
            match = re.search(pattern, text)
            return match.group(1).strip() if match else ""

        surat_jalan = get_val(r'^\s*([^\n]+)', doc)
        trip_name = get_val(r'Nama Line Haul Trip:\s*(.*)', doc)
        waktu_segel = get_val(r'Waktu Segel\s*:\s*([\d\/\s:]+)', doc)
        kode_segel = get_val(r'Kode Segel\s*:\s*(\d+)', doc)
        origin = get_val(r'Origin\s*:\s*([^\n]+)', doc)
        ata = get_val(r'ATA \(Actual Kedatangan\)\s*:\s*([\d\/\s:]+)', doc)
        std = get_val(r'STD \(Jadwal Keberangkatan\)\s*:\s*([\d\/\s:]+)', doc)
        destination = get_val(r'Destination\s*:\s*([^\n]+)', doc)
        sta = get_val(r'STA \(Jadwal Kedatangan\)\s*:\s*([\d\/\s:]+)', doc)
        driver = get_val(r'Nama Driver\s*:\s*([^\n]+)', doc)
        nopol = get_val(r'Nomor Polisi\s*:\s*([^\n]+)', doc)

        raw_vendor = get_val(r'Nama Vendor\s*:\s*([^\n]+)', doc)
        vendor_match = re.match(r'^(.*?)\s+(\d+)\s+(\d+)\s+(\d+)\s+([\d\.]+)\s+(\d+)$', raw_vendor)
        if vendor_match:
            vendor = vendor_match.group(1).strip()
            jml_to, jml_hv, jml_paket = int(vendor_match.group(2)), int(vendor_match.group(3)), int(vendor_match.group(4))
            total_berat, jml_dg = float(vendor_match.group(5)), int(vendor_match.group(6))
        else:
            vendor = raw_vendor
            jml_to = int(get_val(r'Jumlah TO\s*(\d+)', doc) or 0)
            jml_hv = int(get_val(r'Jumlah TO HV\s*(\d+)', doc) or 0)
            jml_paket = int(get_val(r'Jumlah Paket\s*(\d+)', doc) or 0)
            total_berat = float(get_val(r'Total Berat \(kg\)\s*([\d\.]+)', doc) or 0.0)
            jml_dg = int(get_val(r'Jumlah TO DG\s*(\d+)', doc) or 0)

        trips_summary.append({
            "Surat Jalan": surat_jalan, "Destination": destination, "Trip Name": trip_name,
            "Waktu Segel": waktu_segel, "Kode Segel": kode_segel, "Origin": origin,
            "ATA": ata, "STD": std, "STA": sta, "Driver": driver, "Nopol": nopol,
            "Vendor": vendor, "Jml TO": jml_to, "Jml HV": jml_hv, "Jml Paket": jml_paket,
            "Total Berat (kg)": total_berat, "Jml DG": jml_dg
        })

        table_lines = doc.split("\n")
        in_table = False
        for line in table_lines:
            if "#" in line and "Nomor TO" in line:
                in_table = True
                continue
            if in_table:
                if "PIC Gudang" in line or "Notes:" in line: break
                parts = line.split()
                if len(parts) >= 6 and parts[0].isdigit() and parts[1].startswith("TO"):
                    to_details.append({
                        "Surat Jalan": surat_jalan, "Destination": destination, "Nomor TO": parts[1],
                        "Jmlh": int(parts[2]), "Berat (kg)": float(parts[3]),
                        "HV": parts[-3] if len(parts) >= 7 else "N",
                        "TO Type": parts[-2] if len(parts) >= 7 else parts[-1],
                        "DG Type": parts[-1] if ("DG" in parts[-1] or parts[-1] == "Non") else ""
                    })

    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color="1F497D", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    zebra_fill = PatternFill(start_color="F2F5F9", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", fill_type="solid")
    total_fill = PatternFill(start_color="D9E1F2", fill_type="solid")
    thin_border = Border(left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"), top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"))
    double_bottom_border = Border(top=Side(style="thin", color="D9D9D9"), bottom=Side(style="double", color="1F497D"))

    # Sheet 1: Ringkasan Trip
    ws1 = wb.active
    ws1.title = "Ringkasan Trip"
    headers1 = ["No", "No Surat Jalan", "Destination", "Nama Line Haul Trip", "Waktu Segel", "Kode Segel", "Origin", "ATA (Actual Kedatangan)", "STD (Jadwal Keberangkatan)", "STA (Jadwal Kedatangan)", "Nama Driver", "Nomor Polisi", "Nama Vendor", "Jumlah TO", "Jumlah HV", "Jumlah Paket", "Total Berat (kg)", "Jumlah DG"]
    ws1.append(headers1)

    for col_num, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col_num)
        cell.fill, cell.font, cell.alignment = header_fill, header_font, Alignment(horizontal="center", vertical="center")

    for idx, d in enumerate(trips_summary, 1):
        row = [idx, d["Surat Jalan"], d["Destination"], d["Trip Name"], d["Waktu Segel"], d["Kode Segel"], d["Origin"], d["ATA"], d["STD"], d["STA"], d["Driver"], d["Nopol"], d["Vendor"], d["Jml TO"], d["Jml HV"], d["Jml Paket"], d["Total Berat (kg)"], d["Jml DG"]]
        ws1.append(row)
        fill = zebra_fill if idx % 2 == 0 else white_fill
        for col_num in range(1, len(headers1) + 1):
            c = ws1.cell(row=idx + 1, column=col_num)
            c.fill, c.border = fill, thin_border
            if col_num in [14, 15, 16, 18]: c.number_format = '#,##0'; c.alignment = Alignment(horizontal="right")
            elif col_num == 17: c.number_format = '#,##0.000'; c.alignment = Alignment(horizontal="right")

    t_row1 = ws1.max_row + 1
    ws1.cell(row=t_row1, column=1, value="Total").font = Font(name="Arial", size=10, bold=True)
    for col_num in range(1, len(headers1) + 1):
        c = ws1.cell(row=t_row1, column=col_num)
        c.fill, c.font, c.border = total_fill, Font(name="Arial", size=10, bold=True), double_bottom_border

    ws1.cell(row=t_row1, column=14, value=f"=SUM(N2:N{t_row1-1})").number_format = '#,##0'
    ws1.cell(row=t_row1, column=15, value=f"=SUM(O2:O{t_row1-1})").number_format = '#,##0'
    ws1.cell(row=t_row1, column=16, value=f"=SUM(P2:P{t_row1-1})").number_format = '#,##0'
    ws1.cell(row=t_row1, column=17, value=f"=SUM(Q2:Q{t_row1-1})").number_format = '#,##0.000'
    ws1.cell(row=t_row1, column=18, value=f"=SUM(R2:R{t_row1-1})").number_format = '#,##0'

    # Sheet 2: Detail TO
    ws2 = wb.create_sheet(title="Detail TO")
    headers2 = ["No", "No Surat Jalan", "Destination", "Nomor TO", "Jumlah Paket", "Berat (kg)", "HV", "TO Type", "DG Type"]
    ws2.append(headers2)

    for col_num, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.fill, cell.font, cell.alignment = header_fill, header_font, Alignment(horizontal="center", vertical="center")

    for idx, d in enumerate(to_details, 1):
        row = [idx, d["Surat Jalan"], d["Destination"], d["Nomor TO"], d["Jmlh"], d["Berat (kg)"], d["HV"], d["TO Type"], d["DG Type"]]
        ws2.append(row)
        fill = zebra_fill if idx % 2 == 0 else white_fill
        for col_num in range(1, len(headers2) + 1):
            c = ws2.cell(row=idx + 1, column=col_num)
            c.fill, c.border = fill, thin_border
            if col_num == 5: c.number_format = '#,##0'; c.alignment = Alignment(horizontal="right")
            elif col_num == 6: c.number_format = '#,##0.000'; c.alignment = Alignment(horizontal="right")

    t_row2 = ws2.max_row + 1
    ws2.cell(row=t_row2, column=1, value="Total").font = Font(name="Arial", size=10, bold=True)
    for col_num in range(1, len(headers2) + 1):
        c = ws2.cell(row=t_row2, column=col_num)
        c.fill, c.font, c.border = total_fill, Font(name="Arial", size=10, bold=True), double_bottom_border

    if len(to_details) > 0:
        ws2.cell(row=t_row2, column=5, value=f"=SUM(E2:E{t_row2-1})").number_format = '#,##0'
        ws2.cell(row=t_row2, column=6, value=f"=SUM(F2:F{t_row2-1})").number_format = '#,##0.000'

    for ws in [ws1, ws2]:
        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(max(len(str(cell.value or '')) for cell in col) + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ==========================================
# MAIN APP EXECUTION
# ==========================================
if check_password():
    st.set_page_config(page_title="SPX Data Processor", page_icon="📦", layout="centered")
    
    with st.sidebar:
        st.write("👤 Status: **Terautentikasi**")
        if st.button("Logout"):
            st.session_state["password_correct"] = False
            st.rerun()

    st.title("📦 SPX Data Processor App")
    st.write("Aplikasi pengolah data **Laporan Scan PDF** secara otomatis.")

    uploaded_pdf = st.file_uploader("Pilih file PDF (.pdf)", type=["pdf"], key="pdf_uploader")
    
    if uploaded_pdf is not None:
        if st.button("Proses PDF", type="primary"):
            with st.spinner("Sedang membaca dan mengekstrak file PDF..."):
                processed_file = process_pdf_scan(uploaded_pdf.read())
                st.success("Ekstraksi PDF berhasil!")
                st.download_button(
                    label="📥 Unduh Hasil Excel PDF Ekstraksi",
                    data=processed_file,
                    file_name=uploaded_pdf.name.replace(".pdf", ".xlsx").replace(".PDF", ".xlsx"),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
